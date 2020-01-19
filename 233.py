import json
import requests
import openpyxl
import time
import random

url = 'http://218.75.197.123:83/app.do'

login = {
    "method": 'authUser',
    "xh": '你的学号',
    "pwd": '你的密码'
}
#login=json.dumps(login)
#print('登陆')
#login_response = requests.post(url=url, data=login)
#print(login_response.text)
#cookie = '''JSESSIONID=DB66C5C05295C77957632EA903656A15; SERVERID=123; td_cookie=2334126630; 787F87F1B39F8205E341E04C747ABAD6'''
#cookie = requests.utils.dict_from_cookiejar(login_response.cookies)
#cookie = str(cookie)
#print(cookie)
head = {
    'token': "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJleHAiOjE1Nzg5NzgyMzEsImF1ZCI6IjE4NDA4MDAwOTIxIn0.SqlUsRmYN15b6bTJu5hJuuSnbGc5yV5Yj3bJvfssZwg",
    'User-Agent': "PostmanRuntime/7.13.0",
    'Accept': "*/*",
    'Cache-Control': "no-cache",
    'Postman-Token': "a3a8acf0-2476-4da8-bb0d-ebb22c8607cd,2e5d3c92-61fd-4716-9555-6ada3551d75f",
    'Host': "218.75.197.123:83",
    'cookie': "JSESSIONID=DB66C5C05295C77957632EA903656A15; SERVERID=122",
    'accept-encoding': "gzip, deflate",
    'Connection': "keep-alive",
    'cache-control': "no-cache"
    }
print('登陆成功')
xh_head = '1840820'
xh_class_k = [str(x).zfill(2) for x in range(1, 14)]
xh_class_k = ['02']
xh_xh_k = [str(x).zfill(2) for x in range(1, 40)]
wb = openpyxl.load_workbook('data2.xlsx')
sheet = wb.active
sheet.cell(row=1, column=1, value='姓名')
sheet.cell(row=1, column=2, value='科目')
sheet.cell(row=1, column=3, value='成绩')
aim_row = sheet.max_row + 1
sheet.title = '计院成绩信息'
temp = 0
xhs = []
while temp!='-1':
    temp = input()
    xhs.append(temp)
xhs.pop()
#for row in xh_class_k:
    #for col in xh_xh_k:
for xh in xhs:
        #xh = xh_head + row + col
        data = {
            'method': 'getCjcx',
            'xh': xh,
            'xqxnid': '2019-2020-1'
        }
        time.sleep(2)
        print('开始获取')
        grade_response = requests.request("GET", url, headers=head, params=data)
        print(grade_response.text)
        #try:
        grade_response = json.loads(grade_response.text)['result']
        print('获得学号：' + xh)
        if grade_response == []:
            continue
        msg = {
            '学号': xh,
            '姓名': grade_response[0]['xm'],
        }
        if grade_response[0]['kcmc'] != '数字逻辑与数字系统':
            continue
        for sub in grade_response:
            msg[sub['kcmc']] = sub['kcmc'] + ':' + sub['zcj']
        mcol = 1
        print(msg)
        for key in msg.keys():
            sheet.cell(row=aim_row, column=mcol, value=msg[key])
            wb.save(filename='data2.xlsx')
            mcol += 1
        aim_row += 1
        print('success')
        #except:
            #print('空学号')

wb.save(filename='data.xlsx')