import openpyxl

a = [
    [1,2,3],
    [4,4,4],
    [5,6,7]
]
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.active
aim_row = sheet.max_row + 1
sheet.title = '计院成绩信息'
for cell in a:
    print(cell)
    col = 1
    for value in cell:
        sheet.cell(row = aim_row, column=col, value=value)
        col+=1
    aim_row += 1
wb.save(filename='data.xlsx')